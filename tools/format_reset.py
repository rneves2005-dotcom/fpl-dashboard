#!/usr/bin/env python3
"""ONE-SHOT format cleanup for Squads_Data.xlsx.

Run only when Rui explicitly approves. Does NOT modify any cell values —
only row formatting and worksheet dimensions.

Steps:
1. Backup the xlsx with timestamp
2. Reset row heights to default (None → Excel default 15) for all 17,274 data rows
3. Force consistent Arial-11 font on all Player-cell values (defensive)
4. Delete phantom empty rows below row 17,274 (1M+ down to actual data)
5. Save
6. Verify: same player count, same Casemiro/Quenda/Zalazar state
"""

from __future__ import annotations
import sys, shutil, time
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

DB_PATH = Path("/Users/ruimiguelneves/Library/Group Containers/UBF8T346G9.OneDriveSyncClientSuite/OneDrive.noindex/OneDrive/Claude/DB/Squads_Data.xlsx")
REAL_MAX_ROW = 17274  # last row with player data per latest verification


def main():
    if not DB_PATH.exists():
        print(f"ERROR: DB not found at {DB_PATH}", file=sys.stderr)
        return 1

    # Step 1: backup
    ts = time.strftime("%Y%m%d_%H%M%S")
    bak = DB_PATH.with_suffix(f".xlsx.bak_before_format_reset_{ts}")
    print(f"Backing up → {bak.name}")
    shutil.copy2(DB_PATH, bak)

    # Step 2: load
    print("Loading workbook…")
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Players"]
    initial_max = ws.max_row
    print(f"  Sheet max_row before: {initial_max:,}")

    # Step 3: ensure uniform row heights for data rows (None = Excel default)
    print("Resetting row heights for data rows…")
    reset_count = 0
    for r in range(2, REAL_MAX_ROW + 1):
        if r in ws.row_dimensions:
            rd = ws.row_dimensions[r]
            if rd.height is not None or rd.customHeight:
                rd.height = None
                rd.customHeight = False
                reset_count += 1
    print(f"  rows with explicit height cleared: {reset_count}")

    # Step 4: ensure consistent font on cells (defensive — they all looked uniform
    # at the file-format level, but force it again just in case Excel caches a
    # different rendering for some rows)
    print("Enforcing Arial 11 on all Players cells…")
    standard_font = Font(name="Arial", size=11, bold=False, italic=False)
    cell_count = 0
    for row in ws.iter_rows(min_row=2, max_row=REAL_MAX_ROW):
        for c in row:
            if c.font.name != "Arial" or c.font.size != 11.0 or c.font.bold or c.font.italic:
                c.font = standard_font
                cell_count += 1
    print(f"  cells re-fonted: {cell_count}")

    # Step 5: delete phantom rows below REAL_MAX_ROW
    if initial_max > REAL_MAX_ROW:
        phantom = initial_max - REAL_MAX_ROW
        print(f"Deleting {phantom:,} phantom rows below row {REAL_MAX_ROW}…")
        ws.delete_rows(REAL_MAX_ROW + 1, phantom)
        print(f"  done; new max_row will be ~{REAL_MAX_ROW}")

    # Step 6: save
    print("Saving…")
    wb.save(DB_PATH)
    wb.close()

    # Step 7: verify
    print("\nVerifying post-save state…")
    wb = openpyxl.load_workbook(DB_PATH, read_only=True, data_only=True)
    ws = wb["Players"]
    real_rows = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[0] is not None)
    # Confirm critical players still in correct clubs
    checks = {"Casemiro": "Inter Miami", "Geovany Quenda": "Chelsea", "Rodrigo Zalazar": "Sporting"}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] in checks:
            expected = checks[row[2]]
            ok = "✓" if row[4] == expected else "✗"
            print(f"  {ok} {row[2]:25} club={row[4]} (expected {expected})")
    print(f"  Player rows with ID: {real_rows:,}")
    print(f"  Sheet max_row: {ws.max_row:,}")
    wb.close()

    print(f"\nDONE — backup at {bak.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
