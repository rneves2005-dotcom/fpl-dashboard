#!/usr/bin/env python3
"""DEPRECATED — DO NOT USE FOR NORMAL TRANSFERS.

Rui owns all edits to Squads_Data.xlsx. Claude should never mutate the DB.
Use tools/sync.py to refresh world_data.json after Rui edits in Excel.
This script remains only as an emergency / reference tool.

------- Original docstring follows -------
Apply a transfer to Squads_Data.xlsx and regenerate world_data.json.

Usage:
    python3 apply_transfer.py --player "Geovany Quenda" --to-club "Chelsea" --to-div "Premier League"
    python3 apply_transfer.py --player "Otamendi" --to-club "TBD" --to-div ""   # leaving (no destination)

Workflow per invocation:
1. Backup Squads_Data.xlsx to .bak_before_<player-slug>_<timestamp>
2. Locate the player row (substring match, case-insensitive on Player column)
3. Update Clubs, Division; set Previous Club = old Clubs value
4. Save xlsx
5. Regenerate /Users/ruimiguelneves/Code/fpl-dashboard/world_data.json
6. Print a one-line summary
"""

from __future__ import annotations
import argparse, json, re, shutil, sys, time
from collections import defaultdict
from pathlib import Path

import openpyxl
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Squads_Data canonical location is OneDrive/Claude/DB/. macOS TCC blocks shell
# access via the user-facing ~/Library/CloudStorage/… path, but OneDrive's
# internal sync cache (below) is fully readable. Auto-prefer .xlsm (after Rui
# enables VBA auto-sort) over .xlsx.
_DB_DIR = Path("/Users/ruimiguelneves/Library/Group Containers/UBF8T346G9.OneDriveSyncClientSuite/OneDrive.noindex/OneDrive/Claude/DB")
# Priority: new consolidated Squads.xlsm → fallback to Squads_Data.xlsm → fallback .xlsx
DB_PATH = next(
    (p for p in [_DB_DIR / "Squads.xlsm", _DB_DIR / "Squads_Data.xlsm", _DB_DIR / "Squads_Data.xlsx"] if p.exists()),
    _DB_DIR / "Squads.xlsm"
)
JSON_OUT = Path("/Users/ruimiguelneves/Code/fpl-dashboard/world_data.json")
# OneDrive mirror — using the internal sync cache path (same TCC reason as DB_PATH above)
JSON_OUT_ONEDRIVE = Path("/Users/ruimiguelneves/Library/Group Containers/UBF8T346G9.OneDriveSyncClientSuite/OneDrive.noindex/OneDrive/Claude/FPL/world_data.json")

# Maps each Division to the host country (Country column = country of the LEAGUE, not player nationality).
# Player nationality lives in the International column and never changes on transfer.
DIVISION_TO_COUNTRY = {
    "Premier League": "England",
    "Championship": "England",
    "League One": "England",
    "League Two": "England",
    "League Three": "England",
    "Premier League 2": "England",
    "La Liga": "Spain",
    "Bundesliga": "Germany",
    "Serie A": "Italy",
    "Ligue 1": "France",
    "Liga Betclic": "Portugal",
    "Segunda Liga": "Portugal",
    "Liga 3": "Portugal",
    "Liga Revelação U23": "Portugal",
    "Campeonato Portugal": "Portugal",
    "Eredivisie League": "Netherlands",
    "Jupiler League": "Belgium",
    "Super Lig": "Turkey",
    "Scotland Premiership": "Scotland",
    "Greece Superliga": "Greece",
    "Russia Premier League": "Russia",
    "Brasileirao": "Brazil",
    "Brasil B": "Brazil",
    "Argentina Superliga": "Argentina",
    "Major League Soccer": "USA",
    "USL": "USA",
    "Liga MX": "Mexico",
    "J-League": "Japan",
    "China Superleague": "China",
}


def slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")


def find_player(ws, name: str, current_club: str | None = None):
    """Find rows where the Player column contains `name` (case-insensitive).
    If current_club is given, also requires the current Clubs cell to match it."""
    needle = name.lower()
    cc = current_club.lower() if current_club else None
    hits = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell_name = row[2].value
        if cell_name and needle in str(cell_name).lower():
            if cc:
                row_club = (row[4].value or "").lower()
                if cc != row_club and cc not in row_club:
                    continue
            hits.append((i, row))
    return hits


def apply_transfer(player: str, to_club, to_div: str, dry_run: bool = False, leaving: bool = False, current_club: str | None = None) -> int:
    if not DB_PATH.exists():
        print(f"ERROR: DB not found at {DB_PATH}", file=sys.stderr)
        return 1

    print(f"Loading {DB_PATH.name}…")
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Players"]

    hits = find_player(ws, player, current_club)
    if not hits:
        print(f"ERROR: No player matches '{player}'", file=sys.stderr)
        return 2
    if len(hits) > 1:
        print(f"Multiple matches for '{player}':")
        for i, row in hits:
            print(f"  row {i}: {row[2].value} | club: {row[4].value} | div: {row[6].value}")
        print(f"Refine the --player argument to be unique.", file=sys.stderr)
        return 3

    row_idx, row = hits[0]
    old_shirt = row[1].value
    old_club = row[4].value
    old_country = row[5].value
    old_div = row[6].value
    old_prev = row[11].value
    full_name = row[2].value

    # Country column = country of the LEAGUE (not player nationality). Derive from new division.
    new_country = None if leaving else (DIVISION_TO_COUNTRY.get(to_div, old_country) if to_div else None)
    if to_div and to_div not in DIVISION_TO_COUNTRY:
        print(f"WARNING: Division '{to_div}' not in DIVISION_TO_COUNTRY map — keeping existing country {old_country!r}", file=sys.stderr)

    print(f"Found  row {row_idx}: {full_name}")
    print(f"  current : shirt={old_shirt!r} club={old_club!r} country={old_country!r} div={old_div!r} prev={old_prev!r}")
    if leaving:
        print(f"  target  : LEAVING — clearing shirt/club/country/division; prev={old_club!r}")
    else:
        print(f"  target  : shirt=None (clear)  club={to_club!r}  country={new_country!r}  div={to_div!r}  prev={old_club!r}")

    if not leaving and old_club == to_club and old_div == to_div:
        print("No change needed — already there.")
        return 0

    if dry_run:
        print("DRY RUN — not writing.")
        return 0

    # Backup
    ts = time.strftime("%Y%m%d_%H%M%S")
    bak_path = DB_PATH.with_suffix(f".xlsx.bak_before_{slug(full_name)}_{ts}")
    print(f"Backing up → {bak_path.name}")
    shutil.copy2(DB_PATH, bak_path)

    # Mutate cells in place
    row[1].value = None                            # Shirt — clear on transfer (unknown until announced)
    row[4].value = None if leaving else to_club    # Clubs
    row[5].value = new_country                     # Country (league host) or None if leaving
    row[6].value = to_div or None                  # Division
    row[11].value = old_club                       # Previous Club ← old current

    print(f"Saving {DB_PATH.name}…")
    wb.save(DB_PATH)
    wb.close()

    # Regenerate world_data.json
    print("Regenerating world_data.json…")
    regenerate_world_json()

    print(f"DONE  {full_name}: {old_club} → {to_club}")
    return 0


def regenerate_world_json():
    """Re-export Squads_Data.xlsx → world_data.json.

    Two source paths:
    - Real divisions (Premier League, La Liga, ...): grouped by Division → Club
    - Virtual divisions for clubs without a tracked Division: grouped as
      "🌐 {Country}" → Club. Lets users pick e.g. "🌐 Saudi Arabia" to see
      Al-Hilal / Al-Nassr / etc.
    """
    wb = openpyxl.load_workbook(DB_PATH, read_only=True, data_only=True)
    ws = wb["Players"]
    data = defaultdict(lambda: defaultdict(list))
    for row in ws.iter_rows(min_row=2, values_only=True):
        _id, shirt, player, age, club, country, division, goals, games, intl, shirt_int, prev = row
        if not club or not player:
            continue
        if division:
            key = division
        elif country:
            key = f"🌐 {country}"
        else:
            continue  # neither a division nor a country — skip
        data[key][club].append({
            "s": shirt, "n": player, "a": age, "c": country, "g": goals, "p": prev
        })
    for div in data:
        for club in data[div]:
            data[div][club].sort(key=lambda x: (x["s"] is None, x["s"] or 999))
    out = {div: dict(clubs) for div, clubs in data.items()}
    payload = json.dumps(out, ensure_ascii=False)
    JSON_OUT.write_text(payload)
    JSON_OUT_ONEDRIVE.write_text(payload)
    print(f"  wrote {JSON_OUT} ({len(payload):,} bytes)")
    print(f"  wrote {JSON_OUT_ONEDRIVE}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--player", required=True, help="Substring match on Player column (must be unique)")
    ap.add_argument("--to-club", help="New current club (e.g. 'Chelsea')")
    ap.add_argument("--to-div", default="", help="New division (e.g. 'Premier League'). Empty for leaving/TBD.")
    ap.add_argument("--leaving", action="store_true", help="Mark player as having left their current club with no known destination. Sets club/division/country/shirt to null; Previous Club ← old club. Player vanishes from world.html until a destination is announced.")
    ap.add_argument("--current-club", help="Disambiguation: when multiple players match by name, restrict to those currently at this club.")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    if args.leaving:
        to_club = None
        to_div = ""
    else:
        if not args.to_club:
            ap.error("--to-club is required unless --leaving is set")
        to_club = args.to_club
        to_div = args.to_div
    sys.exit(apply_transfer(args.player, to_club, to_div, args.dry_run, leaving=args.leaving, current_club=args.current_club))


if __name__ == "__main__":
    main()
