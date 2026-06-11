#!/usr/bin/env python3
"""Hourly transfer poller — feeds Word tracker doc + JSON cache.

Sources:
  - xcancel.com RSS for ~25 X accounts (FabrizioRomano, Ornstein, club official)
  - maisfutebol cronologia RSS (Portuguese transfer timeline)
  - premierleague.com /en/transfers/2026-27/summer (HTML scrape)

Output:
  - /Users/ruimiguelneves/Code/fpl-dashboard/Transfer_Tracker.docx
    (append-only Word doc; new section per run, table of new entries)
  - tools/cache/last_run.json (dedup state per source)

Run hourly via launchd plist `com.fpl.transfer-poll.plist`.
"""

from __future__ import annotations
import os
import re
import sys
import json
import time
import hashlib
from pathlib import Path
from datetime import datetime, timezone
from typing import Any

def title_hash(title: str) -> str:
    """Stable short hash of title for dedup key · catches title freshness."""
    norm = re.sub(r"\s+", " ", (title or "").strip().lower())[:300]
    return hashlib.sha1(norm.encode("utf-8")).hexdigest()[:10]

try:
    import feedparser
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing dep: {e}\nInstall: pip3 install feedparser requests beautifulsoup4 openpyxl", file=sys.stderr)
    sys.exit(1)

# ─── Config ──────────────────────────────────────────────────────────────

ROOT = Path(__file__).resolve().parent.parent
CACHE_DIR = ROOT / "tools" / "cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
LAST_RUN_PATH = CACHE_DIR / "last_run.json"
TRACKER_XLSX = ROOT / "Transfer_Tracker.xlsx"
DESKTOP_TRACKER = Path("/Users/ruimiguelneves/Desktop/Transfer_Tracker.xlsx")
ONEDRIVE_VISIBLE = Path("/Users/ruimiguelneves/Library/CloudStorage/OneDrive-Personal/Claude/FPL/Transfer_Tracker.xlsx")
ONEDRIVE_TRACKER = Path("/Users/ruimiguelneves/Library/Group Containers/UBF8T346G9.OneDriveSyncClientSuite/OneDrive.noindex/OneDrive/Claude/FPL/Transfer_Tracker.xlsx")

XCANCEL = "https://nitter.net"  # primary Nitter mirror · fallback to others if needed
NITTER_FALLBACKS = ["https://nitter.net", "https://xcancel.com", "https://lightbrd.com"]

X_ACCOUNTS = [
    # Journos
    "FabrizioRomano", "David_Ornstein",
    # Official
    "premierleague",
    # Top-6 clubs
    "NUFC", "LFC", "ManUtd", "ManCity", "Arsenal",
    # Top-12
    "SpursOfficial", "ChelseaFC", "CPFC", "Everton", "AVFCOfficial",
    # Mid
    "LCFC", "Wolves", "SouthamptonFC", "OfficialBHAFC", "NFFC",
    # Promoted + relegated
    "LUFC", "HullCity", "IpswichTown", "Coventry_City",
    "WestHam", "BurnleyOfficial",
    # Other PL
    "BrentfordFC", "afcbournemouth", "FulhamFC", "SunderlandAFC",
]

MAISFUTEBOL_RSS = "http://feeds.feedburner.com/iol/maisfutebol"
MAISFUTEBOL_CRONOLOGIA = "https://maisfutebol.iol.pt/cronologia/6373897a0cf2254fb2824fe4"
PL_TRANSFERS_URL = "https://www.premierleague.com/en/transfers/2026-27/summer"

# Keyword filters — case-insensitive substring match
# Filter logic: HIGH requires (CONFIRM_WORD AND TRANSFER_WORD) or (HERE_WE_GO / OFICIAL standalone)
#               LOW requires RUMOR_WORD AND TRANSFER_WORD
# This filters out hashtag/promo noise like "#OFFICIAL Twitter video"

CONFIRM_WORDS = [
    "OFICIAL", "OFICIALMENTE", "confirma", "anuncia", "anunciou", "anunciaram",
    "OFFICIAL:", "officially", "confirmed",
    "we are delighted to announce", "have signed", "we have signed",
    "are delighted to announce", "agreed", "deal", "permanent move",
]

TRANSFER_WORDS = [
    # Portuguese
    "transferência", "contratação", "contrato", "saída", "deixa", "chega",
    "regressa", "empréstimo", "transferiu", "abandona", "novo treinador",
    # English
    "signed", "signs", "joins", "joining", "joined", "leaves", "leaving", "departure",
    "departs", "loan", "loaned", "release", "released", "free transfer", "free agent",
    "contract expires", "contract expiry", "appointed", "step down", "steps down",
    "stepping down", "new head coach", "new manager", "thank you", "farewell",
]

# Standalone triggers — single keywords that ALWAYS qualify regardless of context
STANDALONE_HIGH = [
    "HERE WE GO", "Here we go!", "🚨 OFICIAL", "🚨 OFFICIAL",
]

KEYWORDS_LOW_TRANSFER = TRANSFER_WORDS  # transfer word reqd for rumor too
RUMOR_WORDS = [
    # English
    "rumor", "rumour", "talks", "discussions", "talks with", "in advanced negotiations",
    "verbal agreement", "negotiations underway", "interested in", "close to signing",
    "set to sign", "set to join", "wants to sign",
    # Portuguese
    "interessado em", "interessada em", "procura", "alvo", "alvos",
    "negociações avançadas", "negociação avançada", "quer contratar",
    "acordo verbal", "em conversações", "em conversações com", "perto de",
    "apontado", "apontados", "apontada", "rumor", "rumores",
    "tentar", "pretende", "pretendem", "deverá", "negocia", "negociar",
]

# Anti-keywords: presence reduces confidence (likely social/marketing not transfer)
ANTI_WORDS = [
    "match preview", "matchday", "ticket", "kit launch", "behind the scenes",
    "🎶", "🎤", "podcast", "interview", "feature", "matchcam", "highlight",
    "wallpaper", "buy now", "save", "discount",
]

# Posts matching these trigger link-follow → fetch linked article + extract names
EXPAND_PATTERNS = [
    "retained", "released list", "list of", "list confirmed",
    "retained/released", "releases & retains", "retained and released",
]

# ─── HTTP helpers ────────────────────────────────────────────────────────

USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"

def fetch(url: str, timeout: int = 20) -> str | None:
    try:
        r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=timeout)
        r.raise_for_status()
        return r.text
    except Exception as e:
        print(f"FETCH FAIL {url}: {e}", file=sys.stderr)
        return None

# ─── Source: xcancel.com RSS ─────────────────────────────────────────────

def poll_x_accounts() -> list[dict[str, Any]]:
    """Pull RSS for each X account via Nitter · falls back through mirrors."""
    out = []
    for handle in X_ACCOUNTS:
        body = None
        used_mirror = None
        for mirror in NITTER_FALLBACKS:
            url = f"{mirror}/{handle}/rss"
            body = fetch(url, timeout=15)
            if body and "whitelist" not in body.lower()[:2000] and "<item>" in body:
                used_mirror = mirror
                break
            body = None
        if not body:
            continue
        feed = feedparser.parse(body)
        for e in feed.entries[:20]:  # cap last 20 per account
            entry = {
                "source": f"X:@{handle}",
                "title": e.get("title", "")[:300],
                "summary": e.get("summary", "")[:1500],
                "link": e.get("link", ""),
                "published": e.get("published", ""),
                "id": e.get("id") or e.get("link"),
            }
            out.append(entry)
    return out

# ─── Source: maisfutebol RSS ─────────────────────────────────────────────

def poll_maisfutebol() -> list[dict[str, Any]]:
    """Two sources from maisfutebol:
       1. General RSS feed (broad football news · low transfer signal)
       2. Cronologia transfer-window timeline (h2 > a entries · high signal)
    """
    out = []

    # Source 1: General RSS (kept for breadth)
    body = fetch(MAISFUTEBOL_RSS, timeout=20)
    if body:
        feed = feedparser.parse(body)
        for e in feed.entries[:30]:
            out.append({
                "source": "maisfutebol RSS",
                "title": e.get("title", "")[:300],
                "summary": e.get("summary", "")[:1500],
                "link": e.get("link", ""),
                "published": e.get("published", ""),
                "id": e.get("id") or e.get("link"),
            })

    # Source 2: Cronologia HTML scrape · handles both <h2><a> and <h2 class="titulo">
    cron_html = fetch(MAISFUTEBOL_CRONOLOGIA, timeout=20)
    if cron_html:
        soup = BeautifulSoup(cron_html, "html.parser")
        # Strategy: find all h2 entries · grab title from h2 text · find nearest <a> in parent for the link
        for h2 in soup.find_all("h2"):
            title = h2.get_text(" ", strip=True)
            if not title or len(title) < 10:
                continue
            # Try link in h2 first (some entries have it)
            a = h2.find("a", href=True)
            link = a["href"] if a else None
            if not link:
                # Look in parent / sibling for a link
                parent = h2.find_parent()
                if parent:
                    sibling_a = parent.find("a", href=True)
                    if sibling_a:
                        link = sibling_a["href"]
            if not link:
                # Fall back to cronologia page itself with anchor based on title
                link = f"{MAISFUTEBOL_CRONOLOGIA}#{title[:40]}"
            if not link.startswith("http"):
                link = f"https://maisfutebol.iol.pt{link}"
            # Pull body paragraph for summary (helps filter accuracy)
            body_text = ""
            parent = h2.find_parent()
            if parent:
                p = parent.find_next("p")
                if p:
                    body_text = p.get_text(" ", strip=True)[:500]
            out.append({
                "source": "maisfutebol cronologia",
                "title": title[:300],
                "summary": f"{title} · {body_text}",
                "link": link,
                "published": "",
                "id": f"cron:{link}:{title[:60]}",
            })

    return out

# ─── Source: Premier League HTML ─────────────────────────────────────────

def poll_premier_league() -> list[dict[str, Any]]:
    body = fetch(PL_TRANSFERS_URL, timeout=20)
    if not body:
        return []
    soup = BeautifulSoup(body, "html.parser")
    out = []
    # PL transfers are usually in a list with player + direction
    # Generic crawl — find any link with /transfers/ in href
    for a in soup.find_all("a", href=True):
        txt = a.get_text(strip=True)
        if not txt or len(txt) > 200:
            continue
        href = a["href"]
        if "transfer" in href.lower() or "signing" in txt.lower() or "joins" in txt.lower():
            out.append({
                "source": "PL Official",
                "title": txt,
                "summary": "",
                "link": href if href.startswith("http") else f"https://www.premierleague.com{href}",
                "published": datetime.now(timezone.utc).isoformat(),
                "id": f"PL:{txt}:{href}",
            })
    return out

# ─── Filter ──────────────────────────────────────────────────────────────

def classify_confidence(text: str) -> str:
    """HIGH requires (CONFIRM word AND TRANSFER word) or standalone trigger.
    LOW requires (RUMOR word AND TRANSFER word).
    Anti-words downgrade or skip.
    """
    tl = text.lower()

    # Anti-words = likely social/marketing noise · skip
    for w in ANTI_WORDS:
        if w.lower() in tl:
            return "SKIP"

    # Standalone HIGH triggers (HERE WE GO, 🚨 OFICIAL)
    for w in STANDALONE_HIGH:
        if w.lower() in tl:
            return "HIGH"

    has_transfer = any(w.lower() in tl for w in TRANSFER_WORDS)
    has_confirm = any(w.lower() in tl for w in CONFIRM_WORDS)
    has_rumor = any(w.lower() in tl for w in RUMOR_WORDS)

    if has_confirm and has_transfer:
        return "HIGH"
    if has_rumor and has_transfer:
        return "LOW"
    return "SKIP"

def matched_keywords(text: str) -> list[str]:
    tl = text.lower()
    matched = []
    for w in CONFIRM_WORDS + TRANSFER_WORDS + RUMOR_WORDS + STANDALONE_HIGH:
        if w.lower() in tl:
            matched.append(w)
    return matched[:8]  # cap shown

def should_expand(text: str) -> bool:
    """True if post mentions a 'list' type article that needs link-following."""
    tl = text.lower()
    return any(p.lower() in tl for p in EXPAND_PATTERNS)

def extract_article_url_from_post(summary: str, link: str) -> str | None:
    """Find a club-website link inside the post summary, or fall back to the post link."""
    # Look for canonical club URLs in the summary HTML
    club_patterns = [
        r'https?://(?:www\.)?(?:chelseafc\.com|liverpoolfc\.com|manutd\.com|mancity\.com|arsenal\.com|tottenhamhotspur\.com|cpfc\.co\.uk|everton\.com|bhafc\.com|nffc\.co\.uk|lcfc\.com|burnleyofficial\.com|premierleague\.com|fulhamfc\.com|brentfordfc\.com|afcbournemouth\.com|leedsunited\.com|coventrycity\.com|sunderlandafc\.com|nufc\.co\.uk|whufc\.com|hullcityfc\.com|ipswichtownfc\.com|wolves\.co\.uk|aston-villa\.com)[^\s"\'<>]+',
    ]
    for pat in club_patterns:
        m = re.search(pat, summary or "", flags=re.I)
        if m:
            return m.group(0).rstrip(".,)")
    # Fall back to the post link itself · we'll try following Nitter→Twitter→external link chain
    return None

def extract_player_names_from_article(html: str) -> list[str]:
    """Extract likely player names from club-news HTML.
    Strategy: bullet-list items + headings inside the article body, names of 2-4 capitalized words."""
    soup = BeautifulSoup(html, "html.parser")
    text_blocks = []
    # Common content containers
    for tag in soup.find_all(["li", "p", "h3", "h4", "h2", "strong"]):
        t = tag.get_text(" ", strip=True)
        if 4 < len(t) < 250:
            text_blocks.append(t)
    # Name pattern: 2-4 capitalised words, possibly with apostrophes/hyphens/accents
    name_pat = re.compile(r"\b([A-Z][a-zA-Z'\-]+(?:\s[A-Z][a-zA-Z'\-]+){1,3})\b")
    names = []
    seen = set()
    for blk in text_blocks:
        # Filter out non-name patterns
        if any(skip in blk.lower() for skip in ["read more", "share this", "cookie", "click here", "subscribe"]):
            continue
        for m in name_pat.finditer(blk):
            name = m.group(1).strip()
            # Avoid generic phrases · require 2+ words but reject common false positives
            if name.lower() in {"premier league", "champions league", "europa league", "fa cup", "league cup",
                                "new york", "old trafford", "stamford bridge", "the athletic", "club website",
                                "academy product", "first team", "head coach", "youth team"}:
                continue
            if len(name.split()) < 2:
                continue
            if name not in seen:
                seen.add(name)
                names.append(name)
    return names[:30]  # cap

# ─── Dedup ───────────────────────────────────────────────────────────────

def load_seen() -> set[str]:
    if not LAST_RUN_PATH.exists():
        return set()
    try:
        return set(json.loads(LAST_RUN_PATH.read_text()).get("seen_ids", []))
    except Exception:
        return set()

def save_seen(seen: set[str]):
    # Cap at 5000 to keep file manageable
    if len(seen) > 5000:
        seen = set(list(seen)[-5000:])
    LAST_RUN_PATH.write_text(json.dumps({"seen_ids": sorted(seen), "updated": datetime.now(timezone.utc).isoformat()}, indent=2))

# ─── Excel output (tabs: HIGH · LOW · ABOUT) ────────────────────────────

HEADERS = ["✓", "When (UTC)", "Source", "Title / Headline", "Keywords matched", "Link"]

def ensure_workbook() -> openpyxl.Workbook:
    """Open existing tracker or create new with 3 tabs."""
    if TRACKER_XLSX.exists():
        return openpyxl.load_workbook(str(TRACKER_XLSX))
    wb = openpyxl.Workbook()
    # Default sheet becomes HIGH
    high = wb.active
    high.title = "HIGH"
    low = wb.create_sheet("LOW")
    about = wb.create_sheet("ABOUT")

    for ws, color in ((high, "C6EFCE"), (low, "FFF2CC")):
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.freeze_panes = "A2"
        # Column widths
        widths = {1: 4, 2: 18, 3: 24, 4: 80, 5: 30, 6: 60}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

    # ABOUT tab
    about.cell(row=1, column=1, value="FPL 26/27 Transfer Tracker").font = Font(bold=True, size=14)
    about_rows = [
        "",
        f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "Append-only · ✓ column for marking when applied to Squads_2026.xlsm DB",
        "",
        "TABS:",
        "  HIGH = confirmed transfers (OFICIAL · HERE WE GO · signed · joins · etc)",
        "  LOW = rumored / in-talks tier (likely but unconfirmed)",
        "",
        "SOURCES:",
        "  - X (27 accounts via nitter.net): journalists + clubs",
        "  - maisfutebol.iol.pt RSS",
        "  - premierleague.com official transfers page",
        "",
        "FILTER LOGIC:",
        "  HIGH requires (confirm-word AND transfer-word) OR standalone trigger",
        "  LOW requires (rumor-word AND transfer-word)",
        "  Anti-words (matchday, ticket, kit launch, etc) skip the entry",
        "",
        "SCHEDULE: every hour on the hour via launchd (com.fpl.transfer-poll)",
    ]
    for i, txt in enumerate(about_rows, 2):
        about.cell(row=i, column=1, value=txt)
    about.column_dimensions["A"].width = 100
    return wb

def append_run_entries(wb: openpyxl.Workbook, entries: list[dict[str, Any]]):
    """Append each entry to its appropriate tab (HIGH or LOW)."""
    if not entries:
        return
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    for e in entries:
        tab = "HIGH" if e["confidence"] == "HIGH" else "LOW"
        ws = wb[tab]
        next_row = ws.max_row + 1
        title = re.sub(r"<[^>]+>", "", e["title"])[:300]
        keywords = ", ".join(e["keywords"][:6])
        ws.cell(row=next_row, column=1, value="")  # checkbox
        ws.cell(row=next_row, column=2, value=now_iso)
        ws.cell(row=next_row, column=3, value=e["source"])
        ws.cell(row=next_row, column=4, value=title).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=next_row, column=5, value=keywords).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=next_row, column=6, value=e["link"][:200])
        # Mild row colour for newest run separator
        # (skipping for now to keep file clean on append)

# ─── Main ────────────────────────────────────────────────────────────────

def main():
    started = time.time()
    print(f"[{datetime.now().isoformat()}] poll_transfers START")

    seen = load_seen()

    all_entries = []
    counts = {"X": 0, "maisfutebol": 0, "PL": 0}

    print("Polling X accounts via xcancel…")
    x_entries = poll_x_accounts()
    counts["X"] = len(x_entries)
    all_entries.extend(x_entries)

    print("Polling maisfutebol RSS…")
    mf_entries = poll_maisfutebol()
    counts["maisfutebol"] = len(mf_entries)
    all_entries.extend(mf_entries)

    print("Polling PL Official transfers page…")
    pl_entries = poll_premier_league()
    counts["PL"] = len(pl_entries)
    all_entries.extend(pl_entries)

    print(f"Total raw entries: {len(all_entries)}")

    # Filter + dedup · ID = link + title_hash so titled-edits re-emit
    keep = []
    new_seen = set(seen)
    expand_targets = []  # posts mentioning lists · expand later
    for e in all_entries:
        base_id = e.get("id") or e.get("link") or e.get("title")
        if not base_id:
            continue
        # Compose dedup key with title-hash so updated titles re-emit
        eid = f"{base_id}::{title_hash(e.get('title',''))}"
        e["_dedup_id"] = eid
        if eid in seen:
            continue
        new_seen.add(eid)
        text = f"{e.get('title','')} {e.get('summary','')}"
        conf = classify_confidence(text)
        if conf == "SKIP":
            continue
        e["confidence"] = conf
        e["keywords"] = matched_keywords(text)
        keep.append(e)
        if should_expand(text):
            expand_targets.append(e)

    # Mark expansion-needing posts so user sees them clearly in tracker
    for e in keep:
        text = f"{e.get('title','')} {e.get('summary','')}"
        if should_expand(text):
            e["keywords"] = ["⚠️ FOLLOW-UP CLICK LINK"] + e["keywords"][:5]

    print(f"After filter+dedup: {len(keep)} new entries · {sum(1 for e in keep if '⚠️ FOLLOW-UP CLICK LINK' in e['keywords'])} need follow-up")
    save_seen(new_seen)

    if not keep:
        print(f"No new entries this run · elapsed {time.time()-started:.1f}s")
        return 0

    # Append to Excel tracker (tabs: HIGH · LOW · ABOUT)
    wb = ensure_workbook()
    append_run_entries(wb, keep)
    wb.save(str(TRACKER_XLSX))
    print(f"Wrote {TRACKER_XLSX} ({TRACKER_XLSX.stat().st_size:,} bytes)")

    # Sync to Desktop (most-visible) + OneDrive
    body = TRACKER_XLSX.read_bytes()
    for dest in (DESKTOP_TRACKER, ONEDRIVE_VISIBLE, ONEDRIVE_TRACKER):
        try:
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(body)
            print(f"Synced: {dest}")
        except Exception as e:
            print(f"Sync to {dest} skipped: {e}", file=sys.stderr)

    print(f"Sources: X={counts['X']} maisfutebol={counts['maisfutebol']} PL={counts['PL']}")
    print(f"Elapsed: {time.time()-started:.1f}s")
    return 0

if __name__ == "__main__":
    sys.exit(main())
