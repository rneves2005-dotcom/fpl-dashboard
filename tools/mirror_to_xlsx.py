#!/usr/bin/env python3
"""Copy data from Squads_Data.xlsm → Squads_Data.xlsx (data-only mirror).

Rui owns Squads_Data.xlsm (with VBA auto-sort, dropdowns, formatting).
Squads_2026.xlsm's Power Queries read from Squads_Data.xlsx.
This script keeps the .xlsx mirror in sync with the .xlsm.

NEVER writes to Squads_Data.xlsm. Pure read.
Only writes to Squads_Data.xlsx (the derivative mirror).

Run after Rui edits the .xlsm and tells Claude to refresh. Pivots in
Squads_2026.xlsm will then show fresh data after Data → Refresh All.
"""

from __future__ import annotations
import sys, shutil, time
from pathlib import Path
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

DB_DIR = Path("/Users/ruimiguelneves/Library/Group Containers/UBF8T346G9.OneDriveSyncClientSuite/OneDrive.noindex/OneDrive/Claude/DB")
SRC_XLSM = DB_DIR / "Squads_Data.xlsm"
DST_XLSX = DB_DIR / "Squads_Data.xlsx"
SHEETS_TO_MIRROR = ['Players', 'Clubs', 'Countries', 'Division']


def main():
    if not SRC_XLSM.exists():
        print(f"ERROR: source not found: {SRC_XLSM}", file=sys.stderr)
        return 1

    # Backup the existing .xlsx mirror
    if DST_XLSX.exists():
        ts = time.strftime("%Y%m%d_%H%M%S")
        bak = DST_XLSX.with_suffix(f'.xlsx.bak_mirror_{ts}')
        shutil.copy2(DST_XLSX, bak)
        print(f"Backup existing mirror → {bak.name}")

    print(f"Reading {SRC_XLSM.name}…")
    src_wb = openpyxl.load_workbook(SRC_XLSM, data_only=True, read_only=False)

    print(f"Creating fresh {DST_XLSX.name} (data + tables, no VBA/macros)…")
    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)  # remove default Sheet

    for sheet_name in SHEETS_TO_MIRROR:
        if sheet_name not in src_wb.sheetnames:
            print(f"  ! skipped — '{sheet_name}' not in source")
            continue
        src_ws = src_wb[sheet_name]
        dst_ws = dst_wb.create_sheet(sheet_name)

        # Copy values row by row (skip formatting to keep file lean)
        rows_copied = 0
        max_col = src_ws.max_column
        for row in src_ws.iter_rows(values_only=True):
            if all(v is None for v in row[:max_col]):
                continue  # drop completely empty rows
            dst_ws.append(row[:max_col])
            rows_copied += 1

        # Recreate the Excel Table with the same name so Power Query can find it
        if dst_ws.max_row >= 2:
            last_col_letter = openpyxl.utils.get_column_letter(max_col)
            ref = f"A1:{last_col_letter}{dst_ws.max_row}"
            tbl = Table(displayName=sheet_name, ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False
            )
            dst_ws.add_table(tbl)
        print(f"  ✓ {sheet_name}: {rows_copied} rows")

    src_wb.close()
    dst_wb.save(DST_XLSX)
    dst_wb.close()
    print(f"\nWrote {DST_XLSX}")

    # Spot-check freshness: Faye games
    print("\nFreshness spot-check:")
    wb = openpyxl.load_workbook(DST_XLSX, read_only=True, data_only=True)
    ws = wb['Players']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == 'Souleymane Faye':
            print(f"  Souleymane Faye: games={row[8]} goals={row[7]} (in {DST_XLSX.name})")
            break
    wb.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
